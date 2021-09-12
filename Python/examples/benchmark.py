##############################################################################
#
# Simple Python program to benchmark several Python Excel writing modules.
#
# python bench_excel_writers.py [num_rows] [num_cols]
#
# Copyright 2013-2018, John McNamara, Charlie Clark
#
import sys

try:
    from time import process_time
except ImportError:
    from time import clock as process_time
from random import randint, sample
from string import ascii_letters

import openpyxl
import xlsxwriter
from OpenXLSX import *

# Default to 1 sheet with 1000 rows x 50 cols
# 10 % strings
row_max = 1000000
col_max = 8
sheets = 1
factor = 0.1

if len(sys.argv) > 1:
    row_max = int(sys.argv[1])

if len(sys.argv) > 2:
    col_max = int(sys.argv[2])

if len(sys.argv) > 3:
    sheets = int(sys.argv[3])

if len(sys.argv) > 4:
    factor = float(sys.argv[4])


def random_string():
    """
    Return a random set of letters
    """
    chars = sample(ascii_letters, randint(3, 12))
    return "".join(chars)


strings = []
for r in range(int(sheets * row_max * factor)):
    strings.append([random_string() for c in range(col_max)])


def print_elapsed_time(module_name, elapsed, optimised=False):
    """ Print module run times in a consistent format. """
    if optimised:
        module_name += " (optimised)"
    print("    %-22s: %6.2f" % (module_name, elapsed))


def time_xlsxwriter(optimised=False):
    """ Run XlsxWriter in optimised/constant memory mode. """
    options = {}
    filename = 'xlsxwriter.xlsx'
    module_name = "xlsxwriter"
    if optimised:
        options['constant_memory'] = True
        filename = "xlsxwriter_opt.xlsx"

    start_time = process_time()

    string_rows = iter(strings)
    workbook = xlsxwriter.Workbook(filename,
                                   options=options)
    for r in range(sheets):
        worksheet = workbook.add_worksheet()

        for row in range(row_max):
            if not row % 10:
                data = next(string_rows)
            else:
                data = [row + col for col in range(col_max)]
            worksheet.write_row(row, 0, data)

    workbook.close()

    elapsed = process_time() - start_time
    print_elapsed_time(module_name, elapsed, optimised)
    # os.remove(filename)


def time_openpyxl(optimised=False):
    """ Run OpenPyXL in default mode. """
    module_name = "openpyxl"
    filename = 'openpyxl.xlsx'
    if optimised:
        filename = 'openpyxl_opt.xlsx'

    start_time = process_time()
    string_rows = iter(strings)

    workbook = openpyxl.Workbook(write_only=optimised)
    for r in range(sheets):
        worksheet = workbook.create_sheet()

        for row in range(row_max):
            if not row % 10:
                data = next(string_rows)
            else:
                data = (row + col for col in range(col_max))
            worksheet.append(data)

    # workbook.save(filename)

    elapsed = process_time() - start_time
    print_elapsed_time(module_name, elapsed, optimised)
    # os.remove(filename)


def time_openxlsx():
    module_name = "openxlsx"
    filename = 'openxlsx.xlsx'

    doc = XLDocument()
    doc.create(filename)

    start_time = process_time()
    string_rows = iter(strings)

    workbook = doc.workbook()
    for r in range(sheets):
        worksheet = workbook.worksheet('Sheet1')
        rng = worksheet.range(XLCellReference('A1'), XLCellReference(row_max, col_max))

        for cell in rng:
            if not cell.cellReference().row() % 10:
                cell.value().stringValue = 'test'  # next(string_rows)
            else:
                cell.value().integerValue = (cell.cellReference().row() + cell.cellReference().row() - 1)

    # doc.save()

    elapsed = process_time() - start_time
    print_elapsed_time(module_name, elapsed, False)
    # os.remove(filename)


print("")
print("Versions:")
print("%s: %s" % ('python', sys.version[:5]))
print("%s: %s" % ('openpyxl', openpyxl.__version__))
print("%s: %s" % ('xlsxwriter', xlsxwriter.__version__))
print("")

print("Dimensions:")
print("    Rows = %d" % row_max)
print("    Cols = %d" % col_max)
print("    Sheets = %d" % sheets)
print("    Proportion text = %0.2f" % factor)
print("")

print("Times:")
time_openxlsx()
time_xlsxwriter()
# time_xlsxwriter(optimised=True)
time_openpyxl()
# time_openpyxl(optimised=True)

print("")
